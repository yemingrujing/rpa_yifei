import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from typing import List, Dict, Any, Optional, Union, Callable
import json
import os


class DataHandler:
    def __init__(self):
        self.current_workbook = None
        self.current_sheet = None
        self.current_dataframe = None

    def read_excel(self, file_path: str, sheet_name: Optional[str] = None, 
                   header: int = 0) -> pd.DataFrame:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header)
        self.current_dataframe = df
        return df

    def write_excel(self, file_path: str, data: Union[List[List], pd.DataFrame, Dict],
                   sheet_name: str = 'Sheet1', mode: str = 'write'):
        if isinstance(data, pd.DataFrame):
            if mode == 'append' and os.path.exists(file_path):
                with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                    data.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                data.to_excel(file_path, sheet_name=sheet_name, index=False)
        elif isinstance(data, list):
            df = pd.DataFrame(data)
            if mode == 'append' and os.path.exists(file_path):
                with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                df.to_excel(file_path, sheet_name=sheet_name, index=False)
        elif isinstance(data, dict):
            df = pd.DataFrame([data])
            if mode == 'append' and os.path.exists(file_path):
                with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                df.to_excel(file_path, sheet_name=sheet_name, index=False)

    def append_to_excel(self, file_path: str, data: Union[List[Dict], Dict],
                       sheet_name: str = 'Sheet1'):
        if isinstance(data, dict):
            data = [data]
        
        new_df = pd.DataFrame(data)
        
        if os.path.exists(file_path):
            try:
                existing_df = pd.read_excel(file_path, sheet_name=sheet_name)
                combined_df = pd.concat([existing_df, new_df], ignore_index=True)
                with pd.ExcelWriter(file_path, mode='w') as writer:
                    combined_df.to_excel(writer, sheet_name=sheet_name, index=False)
            except:
                with pd.ExcelWriter(file_path, mode='a') as writer:
                    new_df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            new_df.to_excel(file_path, sheet_name=sheet_name, index=False)

    def read_excel_range(self, file_path: str, sheet_name: str, 
                         start_row: int, end_row: int, 
                         start_col: int, end_col: int) -> List[List]:
        wb = load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]
        
        data = []
        for row in ws.iter_rows(min_row=start_row, max_row=end_row, 
                                min_col=start_col, max_col=end_col):
            data.append([cell.value for cell in row])
        
        return data

    def write_excel_cell(self, file_path: str, sheet_name: str, 
                         row: int, col: int, value: Any):
        wb = load_workbook(file_path)
        ws = wb[sheet_name]
        ws.cell(row=row, column=col, value=value)
        wb.save(file_path)

    def get_excel_sheets(self, file_path: str) -> List[str]:
        wb = load_workbook(file_path, read_only=True)
        return wb.sheetnames

    def create_excel_with_format(self, file_path: str, data: List[Dict], 
                                  headers: Optional[List[str]] = None,
                                  column_widths: Optional[Dict[int, int]] = None):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        
        if not headers and data:
            headers = list(data[0].keys()) if isinstance(data[0], dict) else []
        
        if headers:
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            for row_idx, row_data in enumerate(data, 2):
                if isinstance(row_data, dict):
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))
                else:
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
        
        if column_widths:
            for col, width in column_widths.items():
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        wb.save(file_path)

    def excel_to_json(self, excel_path: str, json_path: str, 
                     sheet_name: Optional[str] = None):
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
        df.to_json(json_path, orient='records', indent=2, force_ascii=False)

    def json_to_excel(self, json_path: str, excel_path: str, 
                     sheet_name: str = 'Sheet1'):
        df = pd.read_json(json_path)
        df.to_excel(excel_path, sheet_name=sheet_name, index=False)

    def filter_dataframe(self, df: pd.DataFrame, conditions: Dict[str, Any]) -> pd.DataFrame:
        filtered = df.copy()
        for column, value in conditions.items():
            if column in filtered.columns:
                if isinstance(value, str) and value.startswith('contains:'):
                    filtered = filtered[filtered[column].astype(str).str.contains(value[9:], na=False)]
                elif isinstance(value, str) and value.startswith('startswith:'):
                    filtered = filtered[filtered[column].astype(str).str.startswith(value[11:], na=False)]
                elif isinstance(value, str) and value.startswith('endswith:'):
                    filtered = filtered[filtered[column].astype(str).str.endswith(value[9:], na=False)]
                elif isinstance(value, (list, tuple)):
                    filtered = filtered[filtered[column].isin(value)]
                else:
                    filtered = filtered[filtered[column] == value]
        return filtered

    def sort_dataframe(self, df: pd.DataFrame, by: Union[str, List[str]], 
                       ascending: Union[bool, List[bool]] = True) -> pd.DataFrame:
        return df.sort_values(by=by, ascending=ascending)

    def group_dataframe(self, df: pd.DataFrame, by: Union[str, List[str]], 
                        agg_func: str = 'sum') -> pd.DataFrame:
        return df.groupby(by).agg(agg_func)

    def merge_dataframes(self, left: pd.DataFrame, right: pd.DataFrame, 
                        on: Union[str, List[str]], how: str = 'inner') -> pd.DataFrame:
        return pd.merge(left, right, on=on, how=how)

    def pivot_dataframe(self, df: pd.DataFrame, index: Union[str, List[str]], 
                       columns: str, values: str, aggfunc: str = 'sum') -> pd.DataFrame:
        return df.pivot_table(index=index, columns=columns, values=values, aggfunc=aggfunc)

    def transform_column(self, df: pd.DataFrame, column: str, 
                         func: Callable[[Any], Any]) -> pd.DataFrame:
        df = df.copy()
        df[column] = df[column].apply(func)
        return df

    def add_calculated_column(self, df: pd.DataFrame, new_column: str, 
                             expression: str) -> pd.DataFrame:
        df = df.copy()
        df[new_column] = df.eval(expression)
        return df

    def read_csv(self, file_path: str, encoding: str = 'utf-8', 
                delimiter: str = ',') -> pd.DataFrame:
        df = pd.read_csv(file_path, encoding=encoding, delimiter=delimiter)
        self.current_dataframe = df
        return df

    def write_csv(self, file_path: str, data: Union[pd.DataFrame, List[Dict]], 
                 encoding: str = 'utf-8', delimiter: str = ','):
        if isinstance(data, pd.DataFrame):
            data.to_csv(file_path, encoding=encoding, sep=delimiter, index=False)
        elif isinstance(data, list):
            df = pd.DataFrame(data)
            df.to_csv(file_path, encoding=encoding, sep=delimiter, index=False)

    def read_json(self, file_path: str, encoding: str = 'utf-8') -> Union[Dict, List]:
        with open(file_path, 'r', encoding=encoding) as f:
            return json.load(f)

    def write_json(self, file_path: str, data: Union[Dict, List], 
                  encoding: str = 'utf-8', indent: int = 2):
        with open(file_path, 'w', encoding=encoding) as f:
            json.dump(data, f, indent=indent, ensure_ascii=False)

    def extract_text_from_pdf(self, pdf_path: str) -> str:
        try:
            import PyPDF2
            text = ""
            with open(pdf_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages:
                    text += page.extract_text() + "\n"
            return text
        except ImportError:
            raise ImportError("PyPDF2 is required for PDF processing. Install with: pip install PyPDF2")
        except Exception as e:
            raise Exception(f"Failed to extract text from PDF: {str(e)}")

    def get_dataframe_info(self, df: pd.DataFrame) -> Dict[str, Any]:
        return {
            'shape': df.shape,
            'columns': list(df.columns),
            'dtypes': df.dtypes.to_dict(),
            'null_counts': df.isnull().sum().to_dict(),
            'head': df.head().to_dict('records')
        }
